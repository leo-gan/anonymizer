"""Excel .xlsx input: load/save, types, formulas, hidden sheets."""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook, load_workbook

from pdf_anonymizer_core.core import anonymize_file, anonymize_tabular_file
from pdf_anonymizer_core.tables import (
    EXCEL_EXTRA_MESSAGE,
    apply_mapping_to_table,
    load_review_text,
    load_table,
    save_table,
)
from pdf_anonymizer_core.utils import save_results


def _build_roster(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws["A1"] = "Name"
    ws["B1"] = "Email"
    ws["C1"] = "Hired"
    ws["D1"] = "EmployeeId"
    ws["E1"] = "Active"
    ws["F1"] = "Notes"
    ws["G1"] = "Ref"
    ws["A2"] = "Jane Doe"
    ws["B2"] = "jane@acme.com"
    ws["C2"] = datetime(2020, 1, 15)
    ws["C2"].number_format = "YYYY-MM-DD"
    ws["D2"] = 123456789
    ws["D2"].number_format = "000-00-0000"
    ws["E2"] = True
    ws["F2"] = "Called John Smith about the Austin office."
    ws["G2"] = "=A2"
    ws.column_dimensions["A"].width = 24
    ws.merge_cells("A3:B3")
    ws["A3"] = "Merged block"

    hidden = wb.create_sheet("Secrets")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "Email"
    hidden["A2"] = "hidden@acme.com"

    wb.save(path)
    wb.close()
    return path


def _cells(path: Path, sheet: str) -> dict[tuple[int, int], object]:
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb[sheet]
        return {
            (cell.row, cell.column): cell
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None and cell.value != ""
        }
    finally:
        wb.close()


def _any_formula(path: Path) -> bool:
    wb = load_workbook(path, data_only=False)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        return True
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        return True
                    if type(value).__name__ in {"ArrayFormula", "DataTableFormula"}:
                        return True
        return False
    finally:
        wb.close()


def _anonymize_xlsx(path: Path, **kwargs):
    defaults = dict(
        characters_to_anonymize=1000,
        prompt_template="unused",
        model_name="unused",
        use_llm=False,
    )
    defaults.update(kwargs)
    return anonymize_tabular_file(str(path), **defaults)


class TestXlsxLoad:
    def test_multi_sheet_includes_hidden(self, tmp_path) -> None:
        src = _build_roster(tmp_path / "roster.xlsx")
        doc = load_table(str(src))
        assert doc.kind == "xlsx"
        assert [sheet.name for sheet in doc.sheets] == ["Employees", "Secrets"]
        assert doc.sheets[0].hidden is False
        assert doc.sheets[1].hidden is True
        lookup = {
            (cell.sheet, cell.row, cell.column): cell
            for sheet in doc.sheets
            for cell in sheet.cells
        }
        assert lookup[("Employees", 2, 2)].kind == "text"
        assert lookup[("Employees", 2, 2)].search_text == "jane@acme.com"
        assert lookup[("Secrets", 2, 1)].search_text == "hidden@acme.com"
        emp_id = lookup[("Employees", 2, 4)]
        assert emp_id.kind == "number"
        assert emp_id.search_text == "123456789"
        assert emp_id.original == 123456789
        hired = lookup[("Employees", 2, 3)]
        assert hired.kind == "date"
        assert hired.search_text == "2020-01-15"
        active = lookup[("Employees", 2, 5)]
        assert active.kind == "bool"
        assert active.search_text == ""
        formula = lookup[("Employees", 2, 7)]
        assert formula.kind == "formula"
        assert formula.formula == "=A2"

    def test_display_format_is_not_search_text(self, tmp_path) -> None:
        src = _build_roster(tmp_path / "roster.xlsx")
        doc = load_table(str(src))
        emp_id = next(
            cell
            for cell in doc.sheets[0].cells
            if cell.row == 2 and cell.column == 4
        )
        assert emp_id.search_text == "123456789"
        assert "123-45-6789" not in emp_id.search_text
        assert emp_id.number_format == "000-00-0000"

    def test_missing_formula_cache_is_empty(self, tmp_path, caplog) -> None:
        src = _build_roster(tmp_path / "roster.xlsx")
        with caplog.at_level("WARNING"):
            doc = load_table(str(src))
        formula = next(
            cell
            for cell in doc.sheets[0].cells
            if cell.kind == "formula"
        )
        assert formula.search_text == ""
        assert formula.original is None
        assert "cached value" in caplog.text


class TestXlsxAnonymize:
    def test_multi_sheet_masks_and_preserves_structure(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        src = _build_roster(tmp_path / "roster.xlsx")
        review, mapping, entity_texts = _anonymize_xlsx(src)
        assert "jane@acme.com" not in review
        assert "hidden@acme.com" not in review
        assert "EMAIL_1" in review
        assert "# Sheet: Employees" in review
        assert "# Sheet: Secrets" in review
        save_results(
            review,
            {written: original for original, written in mapping.items()},
            str(src),
            entity_texts=entity_texts,
        )
        out = Path("data/anonymized/roster.anonymized.xlsx")
        assert out.is_file()
        wb = load_workbook(out)
        try:
            assert wb.sheetnames == ["Employees", "Secrets"]
            assert wb["Secrets"].sheet_state == "hidden"
            assert wb["Employees"].column_dimensions["A"].width == 24
            assert "A3:B3" in wb["Employees"].merged_cells
            assert wb["Employees"]["B2"].value.startswith("EMAIL_")
            assert wb["Secrets"]["A2"].value.startswith("EMAIL_")
        finally:
            wb.close()

    def test_untouched_types_stay_native_replaced_are_str(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        src = _build_roster(tmp_path / "roster.xlsx")
        review, mapping, entity_texts = _anonymize_xlsx(src)
        save_results(
            review,
            {written: original for original, written in mapping.items()},
            str(src),
            entity_texts=entity_texts,
        )
        out = Path("data/anonymized/roster.anonymized.xlsx")
        employees = _cells(out, "Employees")
        email = employees[(2, 2)]
        assert isinstance(email.value, str)
        assert email.value.startswith("EMAIL_")
        emp_id = employees[(2, 4)]
        assert emp_id.value == 123456789
        assert isinstance(emp_id.value, int)
        active = employees[(2, 5)]
        assert active.value is True
        hired = employees[(2, 3)]
        assert isinstance(hired.value, datetime)
        assert hired.value == datetime(2020, 1, 15)

    def test_numeric_id_skips_regex_without_llm(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        src = _build_roster(tmp_path / "roster.xlsx")
        review, mapping, entity_texts = _anonymize_xlsx(src)
        assert 123456789 not in mapping
        assert "123456789" not in mapping
        assert "US_PASSPORT_1" not in review
        assert "DRIVERS_LICENSE_US_1" not in review
        save_results(
            review,
            {written: original for original, written in mapping.items()},
            str(src),
            entity_texts=entity_texts,
        )
        emp_id = _cells(Path("data/anonymized/roster.anonymized.xlsx"), "Employees")[
            (2, 4)
        ]
        assert emp_id.value == 123456789

    def test_formulas_never_emitted(self, tmp_path, monkeypatch) -> None:
        monkeypatch.chdir(tmp_path)
        src = _build_roster(tmp_path / "roster.xlsx")
        review, mapping, entity_texts = _anonymize_xlsx(src)
        save_results(
            review,
            {written: original for original, written in mapping.items()},
            str(src),
            entity_texts=entity_texts,
        )
        out = Path("data/anonymized/roster.anonymized.xlsx")
        assert _any_formula(out) is False
        ref = load_workbook(out)["Employees"]["G2"].value
        assert ref is None or (not isinstance(ref, str) or not ref.startswith("="))

    def test_replaced_formula_writes_cached_string(self, tmp_path) -> None:
        src = _build_roster(tmp_path / "roster.xlsx")
        doc = load_table(str(src))
        formula = next(
            cell
            for cell in doc.sheets[0].cells
            if cell.row == 2 and cell.column == 7
        )
        formula.original = "Jane Doe"
        formula.search_text = "Jane Doe"
        apply_mapping_to_table(doc, {"Jane Doe": "PERSON_1"}, ["Jane Doe"])
        dest = tmp_path / "out.xlsx"
        save_table(doc, str(dest))
        assert _any_formula(dest) is False
        assert load_workbook(dest)["Employees"]["G2"].value == "PERSON_1"
        assert load_workbook(dest)["Employees"]["A2"].value == "PERSON_1"

    def test_notes_free_text_is_replaced(self, tmp_path, mocker) -> None:
        src = _build_roster(tmp_path / "roster.xlsx")
        mocker.patch(
            "pdf_anonymizer_core.core.identify_entities_with_llm",
            return_value=[
                {
                    "text": "John Smith",
                    "type": "PERSON",
                    "base_form": "John Smith",
                },
                {"text": "Austin", "type": "LOCATION", "base_form": "Austin"},
            ],
        )
        review, mapping, entity_texts = _anonymize_xlsx(src, use_llm=True)
        assert "John Smith" not in review
        assert "Austin" not in review
        assert mapping["John Smith"].startswith("PERSON_")
        assert mapping["Austin"].startswith("LOCATION_")
        assert "Called" in review
        assert "office." in review
        dest = tmp_path / "notes.xlsx"
        doc = load_table(str(src))
        apply_mapping_to_table(doc, mapping, entity_texts)
        save_table(doc, str(dest))
        notes = load_workbook(dest)["Employees"]["F2"].value
        assert "John Smith" not in notes
        assert "Austin" not in notes
        assert "Called" in notes


class TestXlsxDispatchAndReview:
    def test_anonymize_file_dispatches_xlsx(self, tmp_path) -> None:
        src = _build_roster(tmp_path / "roster.xlsx")
        review, mapping = anonymize_file(
            str(src), 1000, "unused", "unused", use_llm=False
        )
        assert review is not None
        assert "# Sheet: Employees" in review
        assert mapping is not None
        assert "jane@acme.com" in mapping

    def test_load_review_text_flattens_xlsx(self, tmp_path) -> None:
        src = _build_roster(tmp_path / "roster.xlsx")
        text = load_review_text(str(src))
        assert "# Sheet: Employees" in text
        assert "# Sheet: Secrets" in text
        assert "jane@acme.com" in text
        assert " | " in text

    def test_missing_openpyxl_raises_value_error(
        self, tmp_path, monkeypatch
    ) -> None:
        path = tmp_path / "roster.xlsx"
        path.write_bytes(b"pk")
        monkeypatch.setitem(sys.modules, "openpyxl", None)
        with pytest.raises(ValueError, match=r'pdf-anonymizer-core\[excel\]'):
            load_table(str(path))
        assert EXCEL_EXTRA_MESSAGE == (
            "Excel support requires the extra: "
            'pip install "pdf-anonymizer-core[excel]"'
        )
