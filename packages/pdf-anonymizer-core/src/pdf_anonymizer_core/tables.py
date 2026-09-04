"""Table load/save, per-cell apply, and review flatten for CSV and Excel."""

from __future__ import annotations

import csv
import io
import logging
import os
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, Literal, Optional

from pdf_anonymizer_core import conf
from pdf_anonymizer_core.spans import replace_entities

TABLE_SUFFIXES = frozenset({".csv", ".xlsx"})
REJECT_SPREADSHEET_SUFFIXES = frozenset({".xls", ".xlsm", ".ods", ".xlsb"})

EXCEL_EXTRA_MESSAGE = (
    "Excel support requires the extra: "
    'pip install "pdf-anonymizer-core[excel]"'
)

_REJECT_MESSAGES = {
    ".xls": "Legacy .xls is not supported. Re-save as .xlsx or export CSV.",
    ".xlsb": "Legacy .xls is not supported. Re-save as .xlsx or export CSV.",
    ".xlsm": (
        "Macro-enabled workbooks are not supported (macros can re-derive PII). "
        "Re-save as .xlsx."
    ),
    ".ods": (
        "OpenDocument spreadsheets are not supported in this version. "
        "Export CSV or .xlsx."
    ),
}

REGEX_CELL_KINDS = frozenset({"text", "formula"})
CellKind = Literal["empty", "bool", "number", "date", "text", "formula"]


@dataclass
class TableCell:
    sheet: str
    row: int
    column: int
    search_text: str
    original: Any
    kind: CellKind
    number_format: Optional[str] = None
    formula: Optional[str] = None


@dataclass
class TableSheet:
    name: str
    hidden: bool
    max_row: int
    max_column: int
    cells: list[TableCell] = field(default_factory=list)
    merge_ranges: list[str] = field(default_factory=list)
    # Per-row field counts from the source (0 = a truly empty line).
    row_widths: list[int] = field(default_factory=list)


@dataclass
class TableDocument:
    path: str
    kind: Literal["csv", "xlsx"]
    encoding: str = "utf-8"
    had_bom: bool = False
    dialect: dict = field(default_factory=dict)
    sheets: list[TableSheet] = field(default_factory=list)


def is_tabular_path(path: str) -> bool:
    return Path(path).suffix.lower() in TABLE_SUFFIXES


def is_rejected_spreadsheet(path: str) -> bool:
    return Path(path).suffix.lower() in REJECT_SPREADSHEET_SUFFIXES


def rejected_spreadsheet_error(path: str) -> ValueError:
    suffix = Path(path).suffix.lower()
    message = _REJECT_MESSAGES.get(
        suffix, f"Spreadsheet format {suffix} is not supported."
    )
    return ValueError(message)


def _require_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise ValueError(EXCEL_EXTRA_MESSAGE) from exc


def column_letter(column: int) -> str:
    """1-based column index to Excel-style letters (1 → A, 27 → AA)."""
    if column <= 0:
        raise ValueError("column must be 1-based")
    result = ""
    n = column
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def cell_to_search_text(value: Any) -> tuple[str, CellKind]:
    """Serialize a native cell value for NER / deny-list / apply."""
    if value is None or value == "":
        return "", "empty"
    if isinstance(value, bool):
        return "", "bool"
    if isinstance(value, int):
        return str(value), "number"
    if isinstance(value, float):
        if value.is_integer() and abs(value) < 1e15:
            return str(int(value)), "number"
        return format(value, "g"), "number"
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat(), "date"
        return value.isoformat(sep=" ", timespec="seconds"), "date"
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat(), "date"
    if isinstance(value, time):
        return value.isoformat(timespec="seconds"), "date"
    return str(value), "text"


def iter_cells(doc: TableDocument) -> Iterable[TableCell]:
    for sheet in doc.sheets:
        yield from sheet.cells


def _cells_by_address(sheet: TableSheet) -> Dict[tuple[int, int], TableCell]:
    return {(cell.row, cell.column): cell for cell in sheet.cells}


def _dialect_kwargs(dialect: Any) -> dict:
    return {
        "delimiter": dialect.delimiter,
        "quotechar": dialect.quotechar,
        "escapechar": getattr(dialect, "escapechar", None),
        "doublequote": dialect.doublequote,
        "skipinitialspace": dialect.skipinitialspace,
        "quoting": dialect.quoting,
        "lineterminator": getattr(dialect, "lineterminator", "\r\n"),
    }


def _detect_lineterminator(text: str) -> str:
    if "\r\n" in text:
        return "\r\n"
    if "\r" in text:
        return "\r"
    return "\n"


def _decode_csv_bytes(raw: bytes) -> tuple[str, str, bool]:
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig"), "utf-8-sig", True
    for encoding in ("utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding), encoding, False
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1"), "latin-1", False


def _classify_csv_field(raw: str) -> tuple[str, CellKind]:
    if raw == "":
        return "", "empty"
    # Only leading '=' is formula-like. '+' is an E.164 phone prefix.
    if raw.startswith("="):
        return raw, "formula"
    return raw, "text"


def _sniff_dialect(text: str) -> Any:
    sample = text[:8192]
    if not sample.strip():
        return csv.excel
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def load_csv(path: str) -> TableDocument:
    file_size = os.path.getsize(path)
    if file_size > conf.MAX_TABLE_BYTES:
        raise ValueError(
            f"Table file exceeds the size limit of {conf.MAX_TABLE_BYTES} bytes."
        )

    raw = Path(path).read_bytes()
    text, encoding, had_bom = _decode_csv_bytes(raw)
    dialect = _sniff_dialect(text)
    dialect_kwargs = _dialect_kwargs(dialect)
    dialect_kwargs["lineterminator"] = _detect_lineterminator(text)

    try:
        rows = list(csv.reader(io.StringIO(text), dialect))
    except csv.Error as exc:
        raise ValueError(f"CSV is not readable: {exc}") from exc
    max_row = len(rows)
    max_column = max((len(row) for row in rows), default=0)
    row_widths = [len(row) for row in rows]

    sheet = TableSheet(
        name="Sheet1",
        hidden=False,
        max_row=max_row,
        max_column=max_column,
        row_widths=row_widths,
    )
    nonempty = 0
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            search_text, kind = _classify_csv_field(value)
            if kind != "empty":
                nonempty += 1
                if nonempty > conf.MAX_TABLE_CELLS:
                    raise ValueError(
                        "Table has more than "
                        f"{conf.MAX_TABLE_CELLS} non-empty cells."
                    )
            if kind == "empty":
                continue
            sheet.cells.append(
                TableCell(
                    sheet=sheet.name,
                    row=row_idx,
                    column=col_idx,
                    search_text=search_text,
                    original=value,
                    kind=kind,
                    formula=value if kind == "formula" else None,
                )
            )

    return TableDocument(
        path=path,
        kind="csv",
        encoding=encoding,
        had_bom=had_bom,
        dialect=dialect_kwargs,
        sheets=[sheet],
    )


def _check_table_file_size(path: str) -> None:
    file_size = os.path.getsize(path)
    if file_size > conf.MAX_TABLE_BYTES:
        raise ValueError(
            f"Table file exceeds the size limit of {conf.MAX_TABLE_BYTES} bytes."
        )


def _styled_is_formula(value: Any, data_type: Optional[str] = None) -> bool:
    if data_type == "f":
        return True
    if isinstance(value, str) and value.startswith("="):
        return True
    return type(value).__name__ in {"ArrayFormula", "DataTableFormula"}


def _formula_text(value: Any) -> str:
    if isinstance(value, str):
        return value
    text = getattr(value, "text", None)
    if isinstance(text, str):
        return text
    return str(value) if value is not None else ""


def _excel_cell_is_formula(cell: Any) -> bool:
    return _styled_is_formula(cell.value, getattr(cell, "data_type", None))


def _xlsx_cell_replaced(cell: TableCell) -> bool:
    initial, _kind = cell_to_search_text(cell.original)
    return cell.search_text != initial


def _xlsx_assign_value(excel_cell: Any, value: Any) -> None:
    # openpyxl treats a leading '=' as a formula; force a string type so a
    # cached "=A1" cannot re-derive a replaced name.
    excel_cell.value = value
    if isinstance(value, str) and value.startswith("="):
        excel_cell.data_type = "s"


def _open_xlsx_workbooks(path: str) -> tuple[Any, Any]:
    from openpyxl import load_workbook

    styled = None
    try:
        # Two live workbooks + TableDocument is peak RAM; save_results reloads
        # then save_xlsx opens styled again (five loads per run).
        styled = load_workbook(path, data_only=False, read_only=False, keep_vba=False)
        values = load_workbook(path, data_only=True, read_only=True, keep_vba=False)
    except ValueError:
        if styled is not None:
            styled.close()
        raise
    except Exception as exc:
        if styled is not None:
            styled.close()
        raise ValueError(f"Cannot open workbook {path}") from exc
    return styled, values


def _cached_values_map(values_ws: Any) -> Dict[tuple[int, int], Any]:
    cached: Dict[tuple[int, int], Any] = {}
    if values_ws is None:
        return cached
    for row in values_ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != "":
                cached[(cell.row, cell.column)] = cell.value
    return cached


def load_xlsx(path: str) -> TableDocument:
    """Load .xlsx via openpyxl.

    Comments, headers/footers, validation lists, defined names, hyperlinks,
    charts, and pivot caches are not walked and may retain identifiers.
    """
    _require_openpyxl()
    _check_table_file_size(path)

    styled, values = _open_xlsx_workbooks(path)
    missing_cache = 0
    nonempty = 0
    sheets: list[TableSheet] = []
    try:
        values_by_name = {ws.title: ws for ws in values.worksheets}
        # worksheets skips chartsheets; includes hidden / veryHidden.
        for ws in styled.worksheets:
            cached = _cached_values_map(values_by_name.get(ws.title))
            merge_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            merge_origins = {
                (rng.min_row, rng.min_col) for rng in ws.merged_cells.ranges
            }
            max_row = ws.max_row or 0
            max_column = ws.max_column or 0
            sheet = TableSheet(
                name=ws.title,
                hidden=ws.sheet_state != "visible",
                max_row=max_row,
                max_column=max_column,
                merge_ranges=merge_ranges,
            )
            if max_row and max_column:
                for row in ws.iter_rows(
                    min_row=1, max_row=max_row, max_col=max_column
                ):
                    for scell in row:
                        addr = (scell.row, scell.column)
                        styled_value = scell.value
                        cached_value = cached.get(addr)
                        is_formula = _styled_is_formula(
                            styled_value, getattr(scell, "data_type", None)
                        )
                        present = (
                            is_formula
                            or (styled_value is not None and styled_value != "")
                            or cached_value is not None
                            or addr in merge_origins
                        )
                        if not present:
                            continue
                        if is_formula:
                            if cached_value is None:
                                missing_cache += 1
                                search_text = ""
                                original: Any = None
                            else:
                                search_text, _kind = cell_to_search_text(
                                    cached_value
                                )
                                original = cached_value
                            kind: CellKind = "formula"
                            formula = _formula_text(styled_value)
                        else:
                            value = (
                                styled_value
                                if styled_value is not None and styled_value != ""
                                else cached_value
                            )
                            search_text, kind = cell_to_search_text(value)
                            original = value
                            formula = None
                        if kind != "empty":
                            nonempty += 1
                            if nonempty > conf.MAX_TABLE_CELLS:
                                raise ValueError(
                                    "Table has more than "
                                    f"{conf.MAX_TABLE_CELLS} non-empty cells."
                                )
                        elif addr not in merge_origins:
                            continue
                        sheet.cells.append(
                            TableCell(
                                sheet=sheet.name,
                                row=scell.row,
                                column=scell.column,
                                search_text=search_text,
                                original=original,
                                kind=kind,
                                number_format=scell.number_format,
                                formula=formula,
                            )
                        )
            sheets.append(sheet)
    finally:
        styled.close()
        values.close()

    if missing_cache:
        logging.warning(
            "%s formula(s) had no cached value; treating as empty.",
            missing_cache,
        )

    return TableDocument(path=path, kind="xlsx", sheets=sheets)


def load_table(path: str) -> TableDocument:
    """Load a ``.csv`` or ``.xlsx`` file as a ``TableDocument``.

    Raises ``ValueError`` for rejected spreadsheet suffixes, a missing
    ``[excel]`` extra, or a file over the size / cell cap.
    """
    suffix = Path(path).suffix.lower()
    if suffix in REJECT_SPREADSHEET_SUFFIXES:
        raise rejected_spreadsheet_error(path)
    if suffix == ".csv":
        return load_csv(path)
    if suffix == ".xlsx":
        return load_xlsx(path)
    raise ValueError(f"Not a supported table file: {path}")


def _csv_write_value(cell: Optional[TableCell]) -> Any:
    if cell is None or cell.kind == "empty":
        return ""
    if cell.kind == "bool":
        return cell.original
    if cell.kind == "formula":
        # Neutralize so spreadsheet apps will not evaluate leftover formulas.
        text = cell.search_text if cell.search_text != "" else str(cell.original)
        if not str(text).startswith("'"):
            return "'" + str(text)
        return text
    if cell.search_text != "":
        return cell.search_text
    return "" if cell.original is None else cell.original


def _row_width(sheet: TableSheet, row: int) -> int:
    if sheet.row_widths and 0 <= row - 1 < len(sheet.row_widths):
        return sheet.row_widths[row - 1]
    return sheet.max_column


def save_csv(doc: TableDocument, path: str) -> None:
    sheet = doc.sheets[0] if doc.sheets else TableSheet(
        name="Sheet1", hidden=False, max_row=0, max_column=0
    )
    lookup = _cells_by_address(sheet)
    encoding = "utf-8-sig" if doc.had_bom else "utf-8"
    dialect = dict(doc.dialect) if doc.dialect else _dialect_kwargs(csv.excel)
    with open(path, "w", encoding=encoding, newline="") as handle:
        writer = csv.writer(handle, **dialect)
        for row in range(1, sheet.max_row + 1):
            width = _row_width(sheet, row)
            if width == 0:
                writer.writerow([])
                continue
            values = [
                _csv_write_value(lookup.get((row, col)))
                for col in range(1, width + 1)
            ]
            writer.writerow(values)


def save_xlsx(doc: TableDocument, path: str) -> None:
    _require_openpyxl()
    from openpyxl import load_workbook

    try:
        styled = load_workbook(doc.path, data_only=False, read_only=False, keep_vba=False)
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Cannot open workbook {doc.path}") from exc

    dropped = 0
    try:
        cells_by_sheet = {sheet.name: _cells_by_address(sheet) for sheet in doc.sheets}
        for ws in styled.worksheets:
            lookup = cells_by_sheet.get(ws.title, {})
            recorded_formulas: set[tuple[int, int]] = set()
            for (row, col), tcell in lookup.items():
                excel_cell = ws.cell(row=row, column=col)
                if tcell.kind == "formula" or _excel_cell_is_formula(excel_cell):
                    dropped += 1
                    recorded_formulas.add((row, col))
                if _xlsx_cell_replaced(tcell):
                    _xlsx_assign_value(excel_cell, tcell.search_text)
                elif tcell.kind == "formula":
                    _xlsx_assign_value(excel_cell, tcell.original)
            max_row = ws.max_row or 0
            max_column = ws.max_column or 0
            if max_row and max_column:
                for row_cells in ws.iter_rows(
                    min_row=1, max_row=max_row, max_col=max_column
                ):
                    for excel_cell in row_cells:
                        addr = (excel_cell.row, excel_cell.column)
                        if addr in recorded_formulas:
                            continue
                        if getattr(excel_cell, "data_type", None) == "f":
                            dropped += 1
                            excel_cell.value = None
        styled.save(path)
    finally:
        styled.close()

    if dropped:
        logging.info("Dropped %s formula(s); wrote cached values.", dropped)


def save_table(doc: TableDocument, path: str) -> None:
    """Write ``doc`` as ``.csv`` or ``.xlsx``. Excel formulas are not persisted."""
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx" or doc.kind == "xlsx":
        save_xlsx(doc, path)
        return
    save_csv(doc, path)


def apply_mapping_to_table(
    doc: TableDocument,
    orig_to_written: Dict[str, str],
    entity_texts: Iterable[str],
) -> TableDocument:
    """Replace detected entity texts in each cell. Does not use mapping keys."""
    texts = [text for text in entity_texts if text]
    if not texts:
        return doc
    for cell in iter_cells(doc):
        if not cell.search_text:
            continue
        new = replace_entities(cell.search_text, texts, orig_to_written)
        if new != cell.search_text:
            cell.search_text = new
    return doc


def _flatten_cell_value(cell: Optional[TableCell], *, anonymized: bool) -> str:
    if cell is None or cell.kind in {"empty", "bool"}:
        return ""
    if anonymized:
        return cell.search_text
    if cell.kind == "formula":
        return str(cell.formula if cell.formula is not None else cell.original or "")
    if cell.original is None:
        return cell.search_text
    return str(cell.original)


def flatten_table_for_review(doc: TableDocument, *, anonymized: bool = True) -> str:
    """Row-wise flatten for verify / risk / consolidate.

    Blank line after the sheet header and after every row, including the last,
    so risk windows do not glue a header or the next sheet onto a data row.
    """
    parts: list[str] = []
    for sheet in doc.sheets:
        parts.append(f"# Sheet: {sheet.name}")
        parts.append("")
        lookup = _cells_by_address(sheet)
        for row in range(1, sheet.max_row + 1):
            values = [
                _flatten_cell_value(lookup.get((row, col)), anonymized=anonymized)
                for col in range(1, sheet.max_column + 1)
            ]
            parts.append(" | ".join(values))
            parts.append("")
    if not parts:
        return ""
    return "\n".join(parts) + "\n"


def load_review_text(path: str) -> str:
    """Load text for ``verify`` / ``report``. Tables and Word files are flattened."""
    if is_rejected_spreadsheet(path):
        raise rejected_spreadsheet_error(path)
    if is_tabular_path(path):
        return flatten_table_for_review(load_table(path), anonymized=True)
    from pdf_anonymizer_core.word import (
        flatten_docx_for_review,
        is_rejected_word,
        is_word_path,
        load_docx,
        rejected_word_error,
    )

    if is_rejected_word(path):
        raise rejected_word_error(path)
    if is_word_path(path):
        return flatten_docx_for_review(load_docx(path), anonymized=True)
    return Path(path).read_text(encoding="utf-8")


def unneutralize_csv_equals(text: str) -> str:
    if text.startswith("'") and text[1:].startswith("="):
        return text[1:]
    return text


def header_labels(sheet: TableSheet) -> Dict[int, str]:
    """First-row values when they look like unique headers; else Col A, Col B."""
    lookup = _cells_by_address(sheet)
    texts: list[str] = []
    for col in range(1, sheet.max_column + 1):
        cell = lookup.get((1, col))
        if cell and cell.kind == "text" and cell.search_text.strip():
            texts.append(cell.search_text.strip())
        else:
            texts.append("")
    if texts and all(texts) and len({item.lower() for item in texts}) == len(texts):
        return {index + 1: texts[index] for index in range(len(texts))}
    return {
        index: f"Col {column_letter(index)}"
        for index in range(1, sheet.max_column + 1)
    }


def write_anonymized_table(
    source_path: str,
    dest_path: str,
    orig_to_written: Dict[str, str],
    entity_texts: Iterable[str],
) -> None:
    doc = load_table(source_path)
    apply_mapping_to_table(doc, orig_to_written, entity_texts)
    dest = Path(dest_path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    save_table(doc, dest_path)


